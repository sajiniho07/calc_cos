import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook

ws = load_workbook("files/sin.xlsx")["Sheet"]
max_row = ws.max_row

data = []
for i in range(1, max_row + 1):
    data.append([ws[f"A{i}"].value, ws[f"B{i}"].value])

data = np.array(data)
x = data[:, 0]
y = data[:, 1]

dy = y[1:] - y[:-1]
dx = x[1:] - x[:-1]

ydot = y.copy()
ydot[1:] = dy / dx

ydot[0] = ydot[1]

# plt.plot(x, y, label="sin")
# plt.plot(x[:-1], ydot, label="cos")
# plt.xlabel("x")
# plt.ylabel("y")
# plt.title("Sin & Cos")
# plt.legend()
# plt.savefig('files/plot.png')

wb_save = Workbook()
ws_save = wb_save.active

i = 0 
for x_i in x:
    ws_save[f"A{i+1}"] = x_i
    ws_save[f"B{i+1}"] = ydot[i]
    i += 1

wb_save.save("files/cos.xlsx")
